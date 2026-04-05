# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SQL Query Exporter – Version 2.02                                       ║
# ║  Modernisierte GUI mit CustomTkinter                                     ║
# ╚══════════════════════════════════════════════════════════════════════════╝
#
# ÄNDERUNGEN gegenüber Version 2.01:
#   - Datenbankverbindungsfelder wurden aus dem Hauptfenster ausgelagert.
#   - Neuer "⚙ Konfiguration"-Button neben dem SegmentedButton öffnet ein
#     separates CTkToplevel-Fenster mit allen Verbindungseinstellungen.
#   - Das Konfigurationsfenster passt seine Felder dynamisch an den gewählten
#     Datenbanktyp an (MSSQL, MSSQL Windows-Auth, SQLite, MySQL, PostgreSQL).
#   - Bei MSSQL: Toggle-Schalter für Windows-Authentifizierung (kein Passwort
#     nötig) vs. SQL-Login (Benutzer + Passwort).
#   - Verbindungsparameter werden beim Speichern in config.json geschrieben
#     (Passwort weiterhin ausgenommen).
#   - Hauptfenster ist jetzt kompakter, da der Verbindungsbereich entfällt.

__version__ = "2.02"

# ─────────────────────────────────────────────────────────────────────────────
# Imports
# ─────────────────────────────────────────────────────────────────────────────

# customtkinter ist die moderne Alternative zu tkinter.
# Installation: pip install customtkinter
import customtkinter as ctk

# tkinter wird weiterhin für filedialog und messagebox benötigt,
# da customtkinter diese Dialoge nicht selbst mitbringt.
import tkinter as tk
from tkinter import filedialog, messagebox

# ttk wird nur noch für den Fortschrittsbalken benötigt
# (customtkinter hat keinen animierten "indeterminate"-Balken).
from tkinter import ttk

import os
import json
from datetime import datetime
from dotenv import load_dotenv
import pandas as pd
from openpyxl.utils import get_column_letter
import threading

# Prüfen ob pyodbc installiert ist.
try:
    import pyodbc
except ImportError:
    raise RuntimeError("❌ Paket 'pyodbc' nicht installiert. Bitte: pip install pyodbc")

# Prüfen ob psycopg2 installiert ist.
try:
    import psycopg2
except ImportError:
    raise RuntimeError("❌ Paket 'psycopg2' nicht installiert. Bitte: pip install psycopg2")

# Prüfen ob chardet installiert ist.
try:
    import chardet
except ImportError:
    raise RuntimeError("❌ Paket 'chardet' nicht installiert. Bitte: pip install chardet")

# load_dotenv() liest die .env-Datei im Projektordner ein und macht alle dort
# definierten Variablen (z.B. DB_PASSWORD) im Programm über os.getenv() verfügbar.
load_dotenv()


# ─────────────────────────────────────────────────────────────────────────────
# CustomTkinter Erscheinungsbild konfigurieren
# ─────────────────────────────────────────────────────────────────────────────
#
# ctk.set_appearance_mode() legt das Farbschema fest:
#   "dark"   → dunkles Design (Standard hier)
#   "light"  → helles Design
#   "system" → folgt automatisch den Windows/macOS-Einstellungen
#
# ctk.set_default_color_theme() wählt das Farb-Theme für Buttons, Slider usw.
#   Verfügbar: "blue" (Standard), "green", "dark-blue"
#   Wir verwenden "dark-blue", da es am besten zum dunklen Lila-Design passt.
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")


# ─────────────────────────────────────────────────────────────────────────────
# Konfiguration (zuletzt genutzte Pfade merken)
# ─────────────────────────────────────────────────────────────────────────────

# CONFIG_FILE speichert den vollständigen Pfad zur config.json-Datei.
# os.path.abspath(__file__) gibt den absoluten Pfad dieser Python-Datei zurück.
# os.path.dirname(...) extrahiert daraus nur den Ordner.
# os.path.join(..., "config.json") hängt den Dateinamen hinten dran.
# Ergebnis: config.json liegt immer im selben Ordner wie dieses Skript.
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")


def config_laden():
    """Liest die gespeicherte Konfiguration. Gibt leeres Dict zurück falls nicht vorhanden."""
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def config_speichern(data: dict):
    """Speichert die Konfiguration als JSON-Datei."""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except PermissionError:
        print(f"Keine Schreibrechte für die Konfigurationsdatei: {CONFIG_FILE}")
    except OSError as e:
        # Voller Datenträger, ungültiger Pfad usw...
        print(f"Fehler beim Speichern der Konfigurationsdatei: {CONFIG_FILE}: {e}")            


# ─────────────────────────────────────────────────────────────────────────────
# Kernfunktionen – Datenbankverbindungen
# ─────────────────────────────────────────────────────────────────────────────
#
# Jede der folgenden Funktionen stellt eine Verbindung zu einem bestimmten
# Datenbanktyp her und gibt ein sogenanntes "Connection-Objekt" zurück.
# Dieses Objekt wird später an pandas übergeben, damit die SQL-Abfrage
# ausgeführt werden kann. Schlägt die Verbindung fehl, wird ein
# verständlicher Fehlertext ausgegeben.
#
# Übersicht der unterstützten Datenbanktypen:
#   - Microsoft SQL Server  → benötigt: pip install pyodbc
#   - SQLite                → bereits in Python eingebaut, keine Installation nötig
#   - PostgreSQL            → benötigt: pip install psycopg2-binary
#   - MySQL                 → benötigt: pip install mysql-connector-python


def verbinde_mssql(server, database, user, password,trusted=False):
    """
    Stellt eine Verbindung zu einem Microsoft SQL Server her.
    trusted=True  → Windows-Authentifizierung (kein Benutzer/Passwort nötig)
    trusted=False → SQL-Login mit Benutzer und Passwort
    """
    
    # Der Connection String ist ein standardisiertes Format, das dem ODBC-Treiber
    # mitteilt, wohin er sich verbinden soll und mit welchen Zugangsdaten.
    
    connection_string = (
    f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};"
    + (
        "Trusted_Connection=yes;"   # trusted=True  → Windows-Auth
        if trusted else
        f"UID={user};PWD={password};" # trusted=False → SQL-Login
        )
    )
    
    try:
        print(f"Versuche Verbindung mit SQL-Datenbank. Der Vorgang kann gegebenenfalls dauern. Server: {server} Datenbank: {database}")
        conn = pyodbc.connect(connection_string, timeout=120)
        print(f"✅ Verbindung zur SQL-Datenbank {database} erfolgreich.")
        return conn
    except Exception as e:
        raise RuntimeError(f"❌ Fehler bei Verbindung zu MS SQL Server: {e}")


def verbinde_sqlite(db_path):
    """
    Stellt eine Verbindung zu einer SQLite-Datenbankdatei her.

    SQLite ist eine dateibasierte Datenbank – es gibt keinen Server.
    Die gesamte Datenbank steckt in einer einzigen .db-Datei auf dem Rechner.
    Das Modul sqlite3 ist in Python bereits eingebaut, es muss nichts
    installiert werden.

    Parameter:
        db_path (str): Vollständiger Dateipfad zur .db-Datei (z.B. C:/daten/meine.db)
    """
    import sqlite3

    # Zuerst prüfen ob die angegebene Datei überhaupt existiert.
    # Falls nicht, lieber sofort eine klare Fehlermeldung ausgeben,
    # statt einen kryptischen sqlite3-Fehler zu werfen.
    if not os.path.isfile(db_path):
        raise RuntimeError(f"❌ SQLite-Datenbankdatei nicht gefunden: {db_path}")
    try:
        conn = sqlite3.connect(db_path)
        return conn
    except Exception as e:
        raise RuntimeError(f"❌ Fehler bei Verbindung zu SQLite: {e}")


def verbinde_postgresql(server, database, user, password, port=5432):
    """
    Stellt eine Verbindung zu einem PostgreSQL-Datenbankserver her.

    PostgreSQL ist ein leistungsfähiges Open-Source-Datenbanksystem.
    Die Verbindung wird über die Bibliothek psycopg2 hergestellt.
    Der Standardport für PostgreSQL ist 5432 – dieser wird automatisch
    vorausgefüllt, kann aber bei Bedarf geändert werden.

    Benötigt: pip install psycopg2-binary

    Parameter:
        server   (str): Adresse oder Name des PostgreSQL-Servers
        database (str): Name der Ziel-Datenbank
        user     (str): Benutzername
        password (str): Passwort
        port     (int): Port des Servers (Standard: 5432)
    """

    try:
        # Anders als bei MSSQL werden die Verbindungsparameter bei psycopg2
        # als einzelne benannte Argumente übergeben – kein langer Connection String.
        conn = psycopg2.connect(
            host=server,
            user=user,
            password=password,
            dbname=database,       # Hinweis: PostgreSQL nennt den Parameter "dbname", nicht "database"
            port=int(port),        # int() stellt sicher, dass der Port als Zahl übergeben wird
            connect_timeout=120    # Verbindungsversuch nach 120 Sekunden abbrechen
        )
        return conn
    except Exception as e:
        raise RuntimeError(f"❌ Fehler bei Verbindung zu PostgreSQL: {e}")


def verbinde_mysql(server, database, user, password, port=3306):
    """
    Stellt eine Verbindung zu einem MySQL-Datenbankserver her.

    MySQL ist eines der weltweit verbreitetsten Datenbanksysteme.
    Die Verbindung wird über mysql-connector-python hergestellt.
    Der Standardport für MySQL ist 3306 – dieser wird automatisch
    vorausgefüllt, kann aber bei Bedarf geändert werden.

    Benötigt: pip install mysql-connector-python

    Parameter:
        server   (str): Adresse oder Name des MySQL-Servers
        database (str): Name der Ziel-Datenbank
        user     (str): Benutzername
        password (str): Passwort
        port     (int): Port des Servers (Standard: 3306)
    """
    # Prüfen ob mysql.connector installiert ist.
    try:
        import mysql.connector
    except ImportError:
        raise RuntimeError(
            "❌ Paket 'mysql-connector-python' nicht installiert.\n"
            "Bitte: pip install mysql-connector-python"
        )
    try:
        conn = mysql.connector.connect(
            host=server,
            user=user,
            password=password,
            database=database,
            port=int(port),
            connection_timeout=120
        )
        return conn
    except Exception as e:
        raise RuntimeError(f"❌ Fehler bei Verbindung zu MySQL: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Kernfunktionen – SQL & Export
# ─────────────────────────────────────────────────────────────────────────────

def lade_sql_datei(file_path):
    """
    Lese die SQL Datei ein und gib den Inhalt zurück.

    Das Encoding (Zeichenkodierung) wird automatisch erkannt, bevor die Datei
    gelesen wird. Das ist notwendig, weil SQL-Dateien je nach Editor in
    unterschiedlichen Kodierungen gespeichert werden können:
        - UTF-8       → moderner Standard (z.B. VS Code, moderne SQL-Editoren)
        - UTF-8-BOM   → UTF-8 mit unsichtbarem Marker am Dateianfang (z.B. SSMS)
        - cp1252      → älterer Windows-Standard (z.B. ältere SQL Server Tools)
        - latin-1     → ähnlich wie cp1252, häufig auf älteren Systemen

    Benötigt: pip install chardet
    """
    try:
        with open(file_path, "rb") as file:
            raw_bytes  = file.read()
            
        # chardet.detect() gibt ein Dict zurück, z.B.:
        # {'encoding': 'utf-8', 'confidence': 0.99, 'language': ''}
        # 'confidence' ist die Sicherheit der Erkennung (0.0 bis 1.0).
        ergebnis   = chardet.detect(raw_bytes )
        encoding   = ergebnis.get("encoding") or "utf-8"  # Fallback auf utf-8
        confidence = ergebnis.get("confidence", 0)    
        
        print(f"📄 Erkanntes Encoding: {encoding} (Sicherheit: {confidence:.0%})")    
        
        # ── Schritt 2: Datei mit erkanntem Encoding als Text einlesen ────
        #
        # errors="replace" → Falls einzelne Zeichen trotz Auto-Detection nicht
        # dekodiert werden können, werden sie durch '?' ersetzt statt einen
        # Absturz zu verursachen. Das ist ein robuster Fallback für Grenzfälle.
        sql_query = raw_bytes.decode(encoding, errors="replace")
            
        print("✅ SQL-Datei erfolgreich eingelesen")
        return sql_query
    except Exception as e:
        raise RuntimeError(f"❌ Fehler bei Einlesen der SQL-Datei: {e}")


def sql_dataframe_erstellen(sql_query, conn):
    """
    Erstelle einen DataFrame aus der SQL Abfrage.

    pandas.read_sql() ist dabei datenbankagnostisch – es funktioniert mit
    jeder Verbindung, egal ob MSSQL, SQLite, PostgreSQL oder MySQL,
    solange ein gültiges Connection-Objekt übergeben wird.
    """
    try:
        df_sql = pd.read_sql(sql_query, conn)
        print("✅ DataFrame erfolgreich erstellt.")
        return df_sql
    except Exception as e:
        raise RuntimeError(f"❌ Fehler beim erstellen des DataFrames: {e}")


def export_to_excel(df_sql, output_path, output_filename):
    """
    Exportiert ein DataFrame in eine Excel-Datei.
    Fügt optional einen Zeitstempel in die letzte Spalte hinzu.

    Parameter:
        df_sql (pd.DataFrame): DataFrame, das exportiert werden soll
        output_path (str): Zielordner
        output_filename (str): Name der Excel-Datei (.xlsx)
    """
    try:
        print("⚙️ Starte Export des DataFrame in Excel.")
        # Vollständigen Pfad erstellen
        full_output_path = os.path.join(output_path, output_filename)

        # Zeitstempel in letzte Spalte einfügen
        print("⚙️ Hinterlege Zeitstempel in die letzte Spalte.")
        df_sql["Export_Zeitstempel"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # ExcelWriter mit openpyxl Engine
        with pd.ExcelWriter(full_output_path, engine="openpyxl") as writer:
            df_sql.to_excel(writer, index=False, sheet_name="Daten")
            # Erste Zeile fixieren
            print("⚙️ Fixiere die erste Zeile.")
            ws = writer.sheets["Daten"]
            ws.freeze_panes = ws["A2"]  # fixiert die erste Zeile

            # Spaltenbreite automatisch anpassen
            print("⚙️ Passe die Spaltenbreite automatisch an")
            for i, col in enumerate(df_sql.columns, 1):
                series = df_sql[col].astype(str)

                try:
                    max_content_length = series.map(len).max()
                    if pd.isna(max_content_length):
                        max_content_length = 0
                except Exception:
                    max_content_length = 0

                max_len = max(int(max_content_length), len(col))
                adjusted_width = min(max_len + 5, 50)
                ws.column_dimensions[get_column_letter(i)].width = adjusted_width

            # AutoFilter auf alle Spaltenüberschriften setzen
            print("⚙️ Setze Autofilter über die erste Zeile.")
            ws.auto_filter.ref = f"A1:{get_column_letter(len(df_sql.columns))}1"

        print(f"✅ DataFrame erfolgreich exportiert nach: {full_output_path}")
        return full_output_path

    except PermissionError as e:
        raise RuntimeError(f"❌ Fehler: Die Datei {output_filename} ist noch geöffnet. Bitte schließen: {e}")
    except Exception as e:
        raise RuntimeError(f"❌ Fehler beim exportieren des DataFrames: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────────────────
#
# NEU: Die Klasse "App" erbt jetzt von ctk.CTk statt von tk.Tk.
# ctk.CTk ist das CustomTkinter-Hauptfenster – es verhält sich genauso wie
# tk.Tk, bringt aber automatisch das moderne Aussehen mit.

class App(ctk.CTk):

    def __init__(self):
        # ctk.CTk initialisieren – das erzeugt das eigentliche Fenster
        # mit dem modernen CustomTkinter-Design.
        super().__init__()

        # Fenstertitel in der Titelleiste
        self.title("SQL Query Exporter")

        # Startgröße des Fensters (Breite x Höhe in Pixeln)
        # Etwas breiter als zuvor, damit der SegmentedButton gut passt.
        self.geometry("740x680")

        # Minimale Fenstergröße festlegen, damit das Layout nicht zusammenbricht
        self.minsize(620, 560)

        # ── Farbpalette ──────────────────────────────────────────────────────
        #
        # CustomTkinter verwendet intern eigene Farbschemata.
        # Für Elemente, die wir manuell einfärben (z.B. den Log, den Progressbar),
        # definieren wir die Farben weiterhin als Variablen.
        #
        # HINWEIS für Einsteiger: Bei ctk-Widgets werden Farben mit den
        # Parametern fg_color (Hintergrund) und text_color (Text) gesetzt –
        # nicht mehr mit bg= und fg= wie bei normalem tkinter.

        BG        = "#1e1e2e"   # Haupt-Hintergrund (sehr dunkel)
        SURFACE   = "#2a2a3e"   # Etwas hellere Fläche (z.B. Log-Hintergrund)
        ACCENT    = "#49658a"   # Akzentfarbe Grau (Buttons, Hervorhebungen)
        FG        = "#cdd6f4"   # Normaler Text (hell auf dunkel)
        FG_DIM    = "#6c7086"   # Abgedunkelter Text (z.B. Labels)
        ENTRY_BG  = "#313244"   # Hintergrund der Eingabefelder
        SUCCESS   = "#a6e3a1"   # Grün für Erfolgsmeldungen im Log
        ERROR     = "#f38ba8"   # Rot für Fehlermeldungen im Log
        WARNING   = "#fab387"   # Orange für Warnmeldungen im Log

        # Hintergrundfarbe des Hauptfensters setzen.
        # Bei ctk.CTk heißt der Parameter "fg_color" (nicht "bg").
        self.configure(fg_color=BG)

        # ── Schriftarten ────────────────────────────────────────────────────
        # Alle verwendeten Schriftarten zentral als Variablen definieren.
        # Bei ctk wird die Schrift als ctk.CTkFont-Objekt übergeben –
        # das ist der ctk-eigene Ersatz für das alte tkinter-Tupel ("Segoe UI", 10).
        FONT_LABEL  = ctk.CTkFont(family="Segoe UI", size=12)
        FONT_BOLD   = ctk.CTkFont(family="Segoe UI", size=12, weight="bold")
        FONT_BTN    = ctk.CTkFont(family="Segoe UI", size=11, weight="bold")
        FONT_TITLE  = ctk.CTkFont(family="Segoe UI", size=15, weight="bold")
        FONT_LOG    = ("Consolas", 9)  # Monospace-Schrift für den Log (bleibt als Tupel,
                                       # da tk.Text kein ctk-Widget ist)

        # ── Farben und Schriften für spätere Methoden merken ─────────────────
        # _verbindungsfelder_aufbauen() wird später auch außerhalb von __init__
        # aufgerufen (z.B. wenn der Benutzer den Datenbanktyp wechselt).
        # Da die Farbvariablen oben nur lokal in __init__ existieren, werden sie
        # hier in einem Dictionary auf der Instanz gespeichert, damit alle
        # Methoden der Klasse darauf zugreifen können.
        self._colors = dict(
            BG=BG, SURFACE=SURFACE, ACCENT=ACCENT, FG=FG,
            FG_DIM=FG_DIM, ENTRY_BG=ENTRY_BG, SUCCESS=SUCCESS,
            ERROR=ERROR, WARNING=WARNING,
            FONT_LABEL=FONT_LABEL, FONT_BTN=FONT_BTN, FONT_LOG=FONT_LOG
        )

        # ── Titelleiste ──────────────────────────────────────────────────────
        # NEU: ctk.CTkFrame ersetzt tk.Frame.
        # corner_radius=0 → keine abgerundeten Ecken (wir wollen einen geraden Balken).
        # fg_color=SURFACE → Hintergrundfarbe des Frames (ctk-Syntax).
        title_bar = ctk.CTkFrame(self, fg_color=SURFACE, corner_radius=0, height=50)
        title_bar.pack(fill="x", padx=0, pady=(0, 0))
        # pack_propagate(False) verhindert, dass der Frame auf seine Kinder-Widgets
        # schrumpft – so bleibt die Höhe von 50px erhalten.
        title_bar.pack_propagate(False)

        # NEU: ctk.CTkLabel ersetzt tk.Label.
        # text_color → Textfarbe (entspricht fg= bei tkinter).
        ctk.CTkLabel(
            title_bar,
            text="⚡  SQL Query Exporter",
            font=FONT_TITLE,
            text_color=ACCENT
        ).pack(side="left", padx=20, pady=10)

        # ── Haupt-Content-Frame ──────────────────────────────────────────────
        # NEU: ctk.CTkFrame mit scrollbarem Inhalt.
        # Dieser Frame enthält alle Eingabeelemente und wächst mit dem Fenster mit.
        # fill="both" + expand=True → füllt den gesamten verbleibenden Platz.
        content = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        content.pack(fill="both", expand=True, padx=20, pady=16)

        # Spalte 1 bekommt weight=1 → sie nimmt übrigen horizontalen Platz ein.
        # Das ist die Basis für das responsive Verhalten der Eingabefelder.
        content.columnconfigure(1, weight=1)

        # ── Gespeicherte Konfiguration laden ────────────────────────────────
        # Beim Start wird config.json eingelesen, um die zuletzt verwendeten
        # Pfade und Einstellungen wiederherzustellen. Falls die Datei noch nicht
        # existiert (erster Start), gibt config_laden() ein leeres Dict zurück
        # und cfg.get(...) liefert jeweils den angegebenen Standardwert.
        cfg = config_laden()

        # ── Abschnitt: Datenbanktyp-Auswahl ─────────────────────────────────
        # Hier wählt der Benutzer, welche Art von Datenbank er verwenden möchte.
        # Der "⚙ Konfiguration"-Button öffnet ein separates Fenster mit den
        # datenbankspezifischen Verbindungseinstellungen.
        ctk.CTkLabel(
            content,
            text="Datenbanktyp",
            font=FONT_LABEL,
            text_color=FG_DIM,
            anchor="w"
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))

        # StringVar hält den aktuell gewählten Datenbanktyp als Text.
        # Beim Start wird der zuletzt gespeicherte Typ aus config.json geladen.
        self.var_db_type = tk.StringVar(value=cfg.get("db_type", "Microsoft SQL Server"))

        # Internes Mapping: Anzeige-Text → interner Wert (für Kompatibilität mit config.json)
        self._db_labels = [
            "Microsoft SQL Server",
            "SQLite",
            "MySQL",
            "PostgreSQL"
        ]
        self._db_label_map = {
            "Microsoft SQL Server":  "Microsoft SQL Server",
            "SQLite":                "SQLite",
            "MySQL":                 "MySQL",
            "PostgreSQL":            "PostgreSQL"
        }
        # Umgekehrtes Mapping: interner Wert → Anzeige-Text
        self._db_label_reverse = {v: k for k, v in self._db_label_map.items()}

        # Aktuellen internen Wert in den Anzeige-Text umrechnen, damit der
        # SegmentedButton beim Start den richtigen Tab markiert.
        aktueller_label = self._db_label_reverse.get(
            cfg.get("db_type", "Microsoft SQL Server"),
            "Microsoft SQL Server"
        )

        # ── Zeile mit SegmentedButton + Konfig-Button ────────────────────────
        # db_row_frame hält SegmentedButton (links, wächst) und Konfig-Button (rechts, fix).
        db_row_frame = ctk.CTkFrame(content, fg_color=BG, corner_radius=0)
        db_row_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 14))
        db_row_frame.columnconfigure(0, weight=1)  # SegmentedButton dehnt sich aus

        self.seg_db_type = ctk.CTkSegmentedButton(
            db_row_frame,
            values=self._db_labels,
            command=self._on_db_type_change,
            fg_color=SURFACE,
            selected_color=ACCENT,
            selected_hover_color="#b4b7d1",
            unselected_color=SURFACE,
            unselected_hover_color=ENTRY_BG,
            text_color=FG,
            font=ctk.CTkFont(family="Segoe UI", size=10)
        )
        self.seg_db_type.set(aktueller_label)
        self.seg_db_type.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        # Konfiguration-Button – öffnet das Verbindungseinstellungs-Fenster
        ctk.CTkButton(
            db_row_frame,
            text="⚙  Konfiguration",
            font=FONT_BTN,
            fg_color=SURFACE,
            text_color=FG,
            hover_color=ENTRY_BG,
            width=148,
            height=32,
            corner_radius=6,
            command=self._konfiguration_oeffnen
        ).grid(row=0, column=1, sticky="e")

        # ── Verbindungsvariablen vorbereiten ─────────────────────────────────
        # Für jedes mögliche Eingabefeld wird eine StringVar angelegt.
        # Reihenfolge der Vorrang-Logik:
        #   1. Wert aus config.json (zuletzt manuell eingegeben)
        #   2. Wert aus .env-Datei  (zentrale Konfigurationsdatei)
        #   3. Leerer String        (Fallback wenn nichts gefunden)
        self.var_server      = tk.StringVar(value=cfg.get("server",   os.getenv("DB_SERVER", "")))
        self.var_database    = tk.StringVar(value=cfg.get("database", os.getenv("DATABASE", "")))
        self.var_user        = tk.StringVar(value=cfg.get("user",     os.getenv("DB_USER", "")))
        self.var_port        = tk.StringVar(value=cfg.get("port",     "3306"))
        self.var_sqlite_path = tk.StringVar(value=cfg.get("sqlite_path", ""))
        # BooleanVar für Windows-Authentifizierung bei MSSQL
        # True  → Windows-Auth (kein Passwort nötig)
        # False → SQL-Login (Benutzer + Passwort)
        self.var_trusted     = tk.BooleanVar(value=cfg.get("trusted", False))

        # Passwort: wird grundsätzlich NICHT in config.json gespeichert.
        # Stattdessen wird das Passwort ausschließlich aus der .env-Datei geladen.
        self.var_password = tk.StringVar(value=os.getenv("DB_PASSWORD", ""))

        # Referenz auf das Konfigurationsfenster (None = geschlossen)
        self._konfig_fenster = None

        # ── Status-Anzeige: Verbindung konfiguriert? ─────────────────────────
        # Zeigt dem Benutzer im Hauptfenster kompakt, welche Verbindung aktiv ist.
        self._lbl_verbindung_status = ctk.CTkLabel(
            content,
            text=self._verbindung_status_text(),
            font=FONT_LABEL,
            text_color=FG_DIM,
            anchor="w"
        )
        self._lbl_verbindung_status.grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 4))

        # ── Trennlinie zwischen Verbindungsbereich und Dateiauswahl ─────────
        ctk.CTkFrame(content, height=1, fg_color=SURFACE, corner_radius=0).grid(
            row=3, column=0, columnspan=2, sticky="ew", pady=(0, 12)
        )

        # ── Abschnitt: SQL-Eingabedatei ──────────────────────────────────────
        # Beschriftung über dem Eingabefeld
        ctk.CTkLabel(
            content,
            text="SQL-Eingabedatei (.sql)",
            font=FONT_LABEL,
            text_color=FG_DIM,
            anchor="w"
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(0, 4))

        # f1 ist ein Unter-Frame, der Eingabefeld und Button nebeneinander enthält.
        f1 = ctk.CTkFrame(content, fg_color=BG, corner_radius=0)
        f1.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        f1.columnconfigure(0, weight=1)

        # StringVar ist eine tkinter-Variable, die automatisch mit dem Eingabefeld verknüpft ist.
        # Ändert sich der Wert programmatisch, aktualisiert sich das Feld sofort – und umgekehrt.
        self.var_input_file = tk.StringVar(value=cfg.get("input_file", ""))

        # NEU: ctk.CTkEntry ersetzt tk.Entry.
        # textvariable → verknüpft das Feld mit der StringVar.
        # fg_color → Hintergrundfarbe des Felds.
        # text_color → Textfarbe.
        # border_color → Rahmenfarbe (sichtbar beim Fokus).
        ctk.CTkEntry(
            f1,
            textvariable=self.var_input_file,
            fg_color=ENTRY_BG,
            text_color=FG,
            border_color=ACCENT,
            font=FONT_LABEL
        ).grid(row=0, column=0, sticky="ew", ipady=2, padx=(0, 8))

        # NEU: ctk.CTkButton ersetzt tk.Button.
        # Alle Hover-Effekte, Rundungen und Farben werden von ctk automatisch
        # verwaltet – wir müssen nur die Grundfarben angeben.
        ctk.CTkButton(
            f1,
            text="Durchsuchen",
            font=FONT_BTN,
            fg_color=SURFACE,
            text_color=ACCENT,
            hover_color=ENTRY_BG,
            width=120,
            command=self.browse_input_file
        ).grid(row=0, column=1, sticky="e")

        # ── Abschnitt: Output-Ordner ──────────────────────────────────────────
        ctk.CTkLabel(
            content,
            text="Output-Ordner",
            font=FONT_LABEL,
            text_color=FG_DIM,
            anchor="w"
        ).grid(row=6, column=0, columnspan=2, sticky="w", pady=(0, 4))

        # f2 ist der Unter-Frame für Eingabefeld + Button des Output-Pfades.
        f2 = ctk.CTkFrame(content, fg_color=BG, corner_radius=0)
        f2.grid(row=7, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        f2.columnconfigure(0, weight=1)

        # Zuletzt genutzten Output-Ordner aus der config.json vorbelegen
        self.var_output_dir = tk.StringVar(value=cfg.get("output_dir", ""))

        # Eingabefeld für den Output-Ordnerpfad
        ctk.CTkEntry(
            f2,
            textvariable=self.var_output_dir,
            fg_color=ENTRY_BG,
            text_color=FG,
            border_color=ACCENT,
            font=FONT_LABEL
        ).grid(row=0, column=0, sticky="ew", ipady=2, padx=(0, 8))

        # "Durchsuchen"-Button für den Ordner-Dialog
        ctk.CTkButton(
            f2,
            text="Durchsuchen",
            font=FONT_BTN,
            fg_color=SURFACE,
            text_color=ACCENT,
            hover_color=ENTRY_BG,
            width=120,
            command=self.browse_output_dir
        ).grid(row=0, column=1, sticky="e")

        # ── Trennlinie ────────────────────────────────────────────────────────
        ctk.CTkFrame(content, height=1, fg_color=SURFACE, corner_radius=0).grid(
            row=8, column=0, columnspan=2, sticky="ew", pady=(4, 14)
        )

        # ── Ausführen-Button ──────────────────────────────────────────────────
        self.btn_run = ctk.CTkButton(
            content,
            text="▶   Ausführen",
            font=FONT_BTN,
            fg_color=ACCENT,
            hover_color="#b4b7d1",
            text_color="white",
            height=42,
            corner_radius=8,
            command=self.run
        )
        self.btn_run.grid(row=9, column=0, columnspan=2, sticky="ew", pady=(0, 14))

        # ── Fortschrittsbalken ────────────────────────────────────────────────
        # Der Fortschrittsbalken zeigt an, dass die Verarbeitung läuft.
        # mode="indeterminate" → animierter Balken ohne genauen Prozentwert
        # (da wir nicht wissen, wie lange die SQL-Abfrage dauert)
        #
        # HINWEIS: CustomTkinter hat zwar ein CTkProgressBar-Widget, dieses
        # unterstützt jedoch keinen "indeterminate"-Modus (d.h. die automatische
        # Hin-und-Her-Animation). Deshalb verwenden wir hier weiterhin das
        # ttk.Progressbar-Widget aus dem normalen tkinter.
        # Um es ans dunkle Design anzupassen, wird ein ttk.Style gesetzt.
        style = ttk.Style(self)
        style.theme_use("clam")  # "clam" ermöglicht freie Farbgestaltung
        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor=SURFACE,
            background=ACCENT,
            bordercolor=SURFACE,
            lightcolor=ACCENT,
            darkcolor=ACCENT
        )

        # Fortschrittsbalken-Widget anlegen und im Grid platzieren
        self.progress = ttk.Progressbar(
            content,
            mode="indeterminate",
            style="Custom.Horizontal.TProgressbar"
        )
        self.progress.grid(row=10, column=0, columnspan=2, sticky="ew", pady=(0, 14))

        # ── Log-Bereich ───────────────────────────────────────────────────────
        ctk.CTkLabel(
            content,
            text="Log",
            font=FONT_BOLD,
            text_color=FG_DIM,
            anchor="w"
        ).grid(row=11, column=0, columnspan=2, sticky="w")

        log_frame = ctk.CTkFrame(content, fg_color=SURFACE, corner_radius=6)
        log_frame.grid(row=12, column=0, columnspan=2, sticky="nsew", pady=(4, 0))

        content.rowconfigure(12, weight=1)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        # Das eigentliche Text-Widget für den Log.
        # HINWEIS: Hier bleibt tk.Text erhalten, weil customtkinter kein
        # eigenes scrollbares, mehrfarbiges Textausgabe-Widget mitbringt.
        # tk.Text kann farbige Tags anzeigen, was für den Log wichtig ist.
        # state="disabled" → verhindert, dass der Benutzer darin tippt.
        # wrap="word" → Zeilenumbruch an Wortgrenzen, nicht mitten im Wort.
        self.log = tk.Text(
            log_frame,
            bg=SURFACE,
            fg=FG,
            font=FONT_LOG,
            relief="flat",
            state="disabled",
            wrap="word",
            insertbackground=FG,
            bd=8,
            pady=4
        )
        self.log.grid(row=0, column=0, sticky="nsew")

        # Scrollbar rechts neben dem Log-Text-Widget.
        # command=self.log.yview → verknüpft die Scrollbar mit dem Text-Widget.
        scrollbar = tk.Scrollbar(
            log_frame,
            command=self.log.yview,
            bg=SURFACE,
            troughcolor=SURFACE,
            activebackground=ACCENT
        )
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Gegenrichtung: Text-Widget informiert die Scrollbar über seine aktuelle Position
        self.log.configure(yscrollcommand=scrollbar.set)

        # ── Farb-Tags für den Log ─────────────────────────────────────────────
        # Tags ermöglichen es, einzelne Zeilen im Text-Widget farbig einzufärben.
        # Beim Schreiben in den Log wird der passende Tag mitgegeben (z.B. "ok", "err").
        self.log.tag_configure("ok",   foreground=SUCCESS)  # Grün  → Erfolg
        self.log.tag_configure("err",  foreground=ERROR)    # Rot   → Fehler
        self.log.tag_configure("warn", foreground=WARNING)  # Orange → Warnung
        self.log.tag_configure("info", foreground=FG_DIM)   # Grau  → normale Info


    # ══════════════════════════════════════════════════════════════════════════
    # Methoden der App-Klasse
    # ══════════════════════════════════════════════════════════════════════════

    def _verbindung_status_text(self):
        """
        Gibt einen kurzen Statustext zurück, der im Hauptfenster anzeigt,
        welche Verbindung aktuell konfiguriert ist.
        """
        db_type = self.var_db_type.get()
        if db_type == "SQLite":
            pfad = self.var_sqlite_path.get()
            if pfad:
                return f"SQLite  ·  {os.path.basename(pfad)}"
            return "SQLite  ·  (keine Datei gewählt)"
        elif db_type == "Microsoft SQL Server":
            server = self.var_server.get() or "—"
            db     = self.var_database.get() or "—"
            auth   = "Windows-Auth" if self.var_trusted.get() else "SQL-Login"
            return f"MSSQL  ·  {server} / {db}  ·  {auth}"
        else:
            server = self.var_server.get() or "—"
            db     = self.var_database.get() or "—"
            return f"{db_type}  ·  {server} / {db}"


    def _on_db_type_change(self, selected_label=None):
        """
        Wird aufgerufen, wenn der Benutzer einen anderen Datenbanktyp auswählt.
        Aktualisiert die interne Variable, den Status-Text und config.json.
        """
        if selected_label:
            interner_wert = self._db_label_map.get(selected_label, selected_label)
            self.var_db_type.set(interner_wert)

        # Port-Standard je nach DB-Typ setzen, wenn das Feld noch leer ist
        db_type = self.var_db_type.get()
        if db_type == "PostgreSQL":
            self.var_port.set(self.var_port.get() or "5432")
        elif db_type == "MySQL":
            self.var_port.set(self.var_port.get() or "3306")

        self._status_aktualisieren()
        self._config_aktualisieren(db_type=self.var_db_type.get())

        # Falls das Konfigurationsfenster gerade offen ist, Inhalt neu aufbauen
        if self._konfig_fenster and self._konfig_fenster.winfo_exists():
            self._konfig_fenster_aufbauen()


    def _status_aktualisieren(self):
        """Aktualisiert den Verbindungs-Statustext im Hauptfenster."""
        self._lbl_verbindung_status.configure(text=self._verbindung_status_text())


    def _konfiguration_oeffnen(self):
        """
        Öffnet das Konfigurationsfenster. Falls es bereits offen ist,
        wird es in den Vordergrund gebracht statt ein zweites zu öffnen.
        """
        if self._konfig_fenster and self._konfig_fenster.winfo_exists():
            self._konfig_fenster.lift()
            self._konfig_fenster.focus()
            return

        c       = self._colors
        BG      = c["BG"]
        SURFACE = c["SURFACE"]
        ACCENT  = c["ACCENT"]
        FONT_TITLE = ctk.CTkFont(family="Segoe UI", size=13, weight="bold")

        # CTkToplevel ist das CustomTkinter-Pendant zu tk.Toplevel.
        # Es erbt automatisch das aktive Farbschema des Hauptfensters.
        win = ctk.CTkToplevel(self)
        win.title("Datenbankverbindung konfigurieren")
        win.geometry("520x480")
        win.minsize(420, 360)
        win.configure(fg_color=BG)
        # Fenster immer im Vordergrund halten (Modal-ähnliches Verhalten)
        win.transient(self)
        win.grab_set()
        self._konfig_fenster = win

        # Titelleiste des Konfigurationsfensters
        title_bar = ctk.CTkFrame(win, fg_color=SURFACE, corner_radius=0, height=44)
        title_bar.pack(fill="x")
        title_bar.pack_propagate(False)
        ctk.CTkLabel(
            title_bar,
            text="⚙  Datenbankverbindung",
            font=FONT_TITLE,
            text_color=ACCENT
        ).pack(side="left", padx=16, pady=10)

        # Inhalt-Frame (wird von _konfig_fenster_aufbauen befüllt)
        self._konfig_content = ctk.CTkFrame(win, fg_color=BG, corner_radius=0)
        self._konfig_content.pack(fill="both", expand=True, padx=20, pady=16)
        self._konfig_content.columnconfigure(1, weight=1)

        self._konfig_fenster_aufbauen()


    def _konfig_fenster_aufbauen(self):
        """
        Befüllt den Inhalt des Konfigurationsfensters passend zum gewählten
        Datenbanktyp. Wird bei jedem Öffnen und bei jedem DB-Typ-Wechsel
        neu aufgerufen.

        Felder je nach Datenbanktyp:
          SQLite          → Dateipfad + Durchsuchen-Button
          MSSQL (SQL)     → Server, Datenbank, Benutzer, Passwort
          MSSQL (Win-Auth)→ Server, Datenbank  (kein Benutzer/Passwort)
          MySQL           → Server, Datenbank, Benutzer, Passwort, Port
          PostgreSQL      → Server, Datenbank, Benutzer, Passwort, Port
        """
        c          = self._colors
        BG         = c["BG"]
        FG         = c["FG"]
        FG_DIM     = c["FG_DIM"]
        ENTRY_BG   = c["ENTRY_BG"]
        ACCENT     = c["ACCENT"]
        SURFACE    = c["SURFACE"]
        FONT_LABEL = c["FONT_LABEL"]
        FONT_BTN   = c["FONT_BTN"]
        FONT_BOLD  = ctk.CTkFont(family="Segoe UI", size=12, weight="bold")

        frame = self._konfig_content

        # Alten Inhalt löschen
        for widget in frame.winfo_children():
            widget.destroy()

        db_type = self.var_db_type.get()

        # ── Hilfsfunktionen ──────────────────────────────────────────────────

        def lbl(text, row, bold=False):
            ctk.CTkLabel(
                frame, text=text,
                font=FONT_BOLD if bold else FONT_LABEL,
                text_color=FG_DIM if not bold else FG,
                anchor="w"
            ).grid(row=row, column=0, columnspan=2, sticky="w", pady=(8, 2))

        def field(label_text, var, row, show=""):
            ctk.CTkLabel(
                frame, text=label_text,
                font=FONT_LABEL, text_color=FG_DIM, anchor="w"
            ).grid(row=row, column=0, sticky="w", pady=(0, 2), padx=(0, 12))
            ctk.CTkEntry(
                frame, textvariable=var,
                fg_color=ENTRY_BG, text_color=FG,
                border_color=ACCENT, font=FONT_LABEL, show=show
            ).grid(row=row, column=1, sticky="ew", ipady=2, pady=(0, 6))

        # ── Felder je nach Datenbanktyp ──────────────────────────────────────

        row = 0

        if db_type == "SQLite":
            lbl("SQLite-Datenbankdatei", row, bold=True); row += 1
            # Dateipfad-Zeile: Eingabefeld + Durchsuchen-Button nebeneinander
            f = ctk.CTkFrame(frame, fg_color=BG, corner_radius=0)
            f.grid(row=row, column=0, columnspan=2, sticky="ew", pady=(0, 10))
            f.columnconfigure(0, weight=1)
            ctk.CTkEntry(
                f, textvariable=self.var_sqlite_path,
                fg_color=ENTRY_BG, text_color=FG,
                border_color=ACCENT, font=FONT_LABEL
            ).grid(row=0, column=0, sticky="ew", ipady=2, padx=(0, 8))
            ctk.CTkButton(
                f, text="Durchsuchen", font=FONT_BTN,
                fg_color=SURFACE, text_color=ACCENT, hover_color=ENTRY_BG,
                width=120, command=self.browse_sqlite_file
            ).grid(row=0, column=1, sticky="e")
            row += 1

        else:
            # ── Server-Verbindungsfelder ─────────────────────────────────────
            lbl("Serververbindung", row, bold=True); row += 1
            field("Server / Host", self.var_server,   row); row += 1
            field("Datenbank",     self.var_database, row); row += 1

            if db_type in ("MySQL", "PostgreSQL"):
                field("Port", self.var_port, row); row += 1

            # ── Authentifizierung ────────────────────────────────────────────
            lbl("Authentifizierung", row, bold=True); row += 1

            if db_type == "Microsoft SQL Server":
                # Toggle: Windows-Auth vs. SQL-Login
                # CTkSwitch zeigt einen modernen Ein/Aus-Schalter.
                switch_frame = ctk.CTkFrame(frame, fg_color=BG, corner_radius=0)
                switch_frame.grid(row=row, column=0, columnspan=2, sticky="ew", pady=(0, 8))
                row += 1

                self._switch_trusted = ctk.CTkSwitch(
                    switch_frame,
                    text="Windows-Authentifizierung (kein Passwort)",
                    variable=self.var_trusted,
                    font=FONT_LABEL,
                    text_color=FG,
                    fg_color=SURFACE,
                    progress_color=ACCENT,
                    command=self._on_trusted_toggle
                )
                self._switch_trusted.pack(anchor="w")

                # Benutzer/Passwort-Felder – werden bei Windows-Auth ausgeblendet
                self._auth_frame = ctk.CTkFrame(frame, fg_color=BG, corner_radius=0)
                self._auth_frame.grid(row=row, column=0, columnspan=2, sticky="ew")
                self._auth_frame.columnconfigure(1, weight=1)
                row += 1

                ctk.CTkLabel(
                    self._auth_frame, text="Benutzer",
                    font=FONT_LABEL, text_color=FG_DIM, anchor="w"
                ).grid(row=0, column=0, sticky="w", pady=(0, 2), padx=(0, 12))
                ctk.CTkEntry(
                    self._auth_frame, textvariable=self.var_user,
                    fg_color=ENTRY_BG, text_color=FG,
                    border_color=ACCENT, font=FONT_LABEL
                ).grid(row=0, column=1, sticky="ew", ipady=2, pady=(0, 6))

                ctk.CTkLabel(
                    self._auth_frame, text="Passwort",
                    font=FONT_LABEL, text_color=FG_DIM, anchor="w"
                ).grid(row=1, column=0, sticky="w", pady=(0, 2), padx=(0, 12))
                ctk.CTkEntry(
                    self._auth_frame, textvariable=self.var_password,
                    fg_color=ENTRY_BG, text_color=FG,
                    border_color=ACCENT, font=FONT_LABEL, show="•"
                ).grid(row=1, column=1, sticky="ew", ipady=2, pady=(0, 6))

                # Sichtbarkeit sofort nach aktuellem Toggle-Zustand setzen
                self._on_trusted_toggle()

            else:
                # MySQL / PostgreSQL: immer Benutzer + Passwort
                field("Benutzer", self.var_user,     row); row += 1
                field("Passwort", self.var_password, row, show="•"); row += 1

        # ── Speichern-Button ─────────────────────────────────────────────────
        ctk.CTkButton(
            frame,
            text="💾  Speichern & Schließen",
            font=FONT_BTN,
            fg_color=ACCENT,
            hover_color="#b4b7d1",
            text_color="white",
            height=38,
            corner_radius=8,
            command=self._konfig_speichern_und_schliessen
        ).grid(row=row + 1, column=0, columnspan=2, sticky="ew", pady=(16, 0))


    def _on_trusted_toggle(self):
        """
        Wird aufgerufen wenn der Windows-Auth-Toggle umgeschaltet wird.
        Blendet die Benutzer/Passwort-Felder ein oder aus.
        """
        if self.var_trusted.get():
            # Windows-Auth aktiv → Benutzer/Passwort-Felder ausblenden
            self._auth_frame.grid_remove()
        else:
            # SQL-Login → Felder wieder einblenden
            self._auth_frame.grid()


    def _konfig_speichern_und_schliessen(self):
        """
        Speichert alle Verbindungseinstellungen in config.json und
        schließt das Konfigurationsfenster. Der Statustext im Hauptfenster
        wird danach aktualisiert.
        """
        self._config_aktualisieren(
            db_type=self.var_db_type.get(),
            server=self.var_server.get().strip(),
            database=self.var_database.get().strip(),
            user=self.var_user.get().strip(),
            port=self.var_port.get().strip(),
            sqlite_path=self.var_sqlite_path.get().strip(),
            trusted=self.var_trusted.get()
            # Passwort wird bewusst NICHT gespeichert
        )
        self._status_aktualisieren()
        if self._konfig_fenster and self._konfig_fenster.winfo_exists():
            self._konfig_fenster.destroy()


    # ────────────────────────────────────────────────────────────────────────
    # Datei-Dialoge
    # ────────────────────────────────────────────────────────────────────────

    def browse_input_file(self):
        """
        Öffnet einen Datei-Auswahl-Dialog für die SQL-Eingabedatei.
        Der Dialog startet im zuletzt genutzten Ordner (aus config.json).
        Nach Auswahl wird der Pfad ins Eingabefeld geschrieben und gespeichert.
        """
        # Startordner ermitteln: Ordner der aktuell eingetragenen Datei verwenden.
        # Falls noch kein Pfad eingetragen ist, wird "/" (Wurzelverzeichnis) genutzt.
        initial = os.path.dirname(self.var_input_file.get()) or "/"

        # Nativen Datei-Dialog öffnen (nur .sql-Dateien werden angezeigt)
        path = filedialog.askopenfilename(
            title="SQL-Datei auswählen",
            initialdir=initial,
            filetypes=[("SQL-Dateien", "*.sql"), ("Alle Dateien", "*.*")]
        )

        # Nur weiterarbeiten, wenn der Benutzer tatsächlich eine Datei ausgewählt hat
        # (und den Dialog nicht abgebrochen hat)
        if path:
            self.var_input_file.set(path)               # Eingabefeld aktualisieren
            self._config_aktualisieren(input_file=path) # Pfad in config.json speichern


    def browse_output_dir(self):
        """
        Öffnet einen Ordner-Auswahl-Dialog für den Output-Ordner.
        Der Dialog startet im zuletzt genutzten Output-Ordner (aus config.json).
        Nach Auswahl wird der Pfad ins Eingabefeld geschrieben und gespeichert.
        """
        # Startordner: aktuell eingetragener Output-Ordner, sonst Wurzelverzeichnis
        initial = self.var_output_dir.get() or "/"

        # Nativen Ordner-Dialog öffnen
        path = filedialog.askdirectory(title="Output-Ordner auswählen", initialdir=initial)

        if path:
            self.var_output_dir.set(path)                # Eingabefeld aktualisieren
            self._config_aktualisieren(output_dir=path)  # Pfad in config.json speichern


    def browse_sqlite_file(self):
        """
        Öffnet einen Datei-Auswahl-Dialog speziell für SQLite-Datenbankdateien.
        Akzeptiert Dateien mit den Endungen .db, .sqlite und .sqlite3 –
        das sind die gängigen Dateiendungen für SQLite-Datenbanken.
        Nach Auswahl wird der Pfad ins Eingabefeld geschrieben und gespeichert.
        """
        initial = os.path.dirname(self.var_sqlite_path.get()) or "/"

        path = filedialog.askopenfilename(
            title="SQLite-Datenbankdatei auswählen",
            initialdir=initial,
            filetypes=[("SQLite-Datenbanken", "*.db *.sqlite *.sqlite3"),
                       ("Alle Dateien", "*.*")]
        )

        if path:
            self.var_sqlite_path.set(path)               # Eingabefeld aktualisieren
            self._config_aktualisieren(sqlite_path=path) # Pfad in config.json speichern


    # ────────────────────────────────────────────────────────────────────────
    # Konfiguration
    # ────────────────────────────────────────────────────────────────────────

    def _config_aktualisieren(self, **kwargs):
        """
        Liest die aktuelle config.json ein, aktualisiert die übergebenen Werte
        und speichert die Datei wieder. So gehen bereits gespeicherte Werte nicht verloren.

        **kwargs bedeutet, dass beliebig viele benannte Parameter übergeben werden können,
        z.B.: self._config_aktualisieren(server="localhost", database="meindb")
        """
        cfg = config_laden()
        cfg.update(kwargs)  # Neue Werte einfügen bzw. bestehende überschreiben
        config_speichern(cfg)


    # ────────────────────────────────────────────────────────────────────────
    # Log
    # ────────────────────────────────────────────────────────────────────────

    def log_write(self, msg, tag="info"):
        """
        Schreibt eine neue Zeile in den Log-Bereich.
        Jede Zeile erhält automatisch einen Zeitstempel (HH:MM:SS).
        Der 'tag' bestimmt die Textfarbe (ok=grün, err=rot, warn=orange, info=grau).

        Das Text-Widget muss zum Schreiben kurz aktiviert (state="normal") und
        danach wieder deaktiviert (state="disabled") werden, damit der Benutzer
        es nicht manuell bearbeiten kann.
        """
        self.log.configure(state="normal")  # Schreibschutz aufheben
        ts = datetime.now().strftime("%H:%M:%S")
        self.log.insert("end", f"[{ts}]  {msg}\n", tag)  # Zeile mit Tag einfügen
        self.log.see("end")                               # Automatisch nach unten scrollen
        self.log.configure(state="disabled")              # Schreibschutz wieder aktivieren


    # ────────────────────────────────────────────────────────────────────────
    # Ausführen
    # ────────────────────────────────────────────────────────────────────────

    def run(self):
        """
        Wird aufgerufen, wenn der Benutzer auf "Ausführen" klickt.
        Prüft zunächst, ob die Eingaben gültig sind.
        Startet die eigentliche Verarbeitung dann in einem separaten Thread,
        damit die GUI während des Exports nicht einfriert.
        """
        # Aktuelle Werte aus den Eingabefeldern auslesen (.strip() entfernt Leerzeichen)
        input_file = self.var_input_file.get().strip()
        output_dir = self.var_output_dir.get().strip()

        # Validierung: Existiert die angegebene SQL-Datei wirklich?
        if not input_file or not os.path.isfile(input_file):
            messagebox.showerror("Fehler", "Bitte eine gültige SQL-Datei auswählen.")
            return

        # Validierung: Existiert der angegebene Output-Ordner wirklich?
        if not output_dir or not os.path.isdir(output_dir):
            messagebox.showerror("Fehler", "Bitte einen gültigen Output-Ordner auswählen.")
            return

        # Verbindungsparameter je nach Datenbanktyp prüfen.
        # SQLite braucht nur eine Datei, alle anderen brauchen mindestens Server + Datenbank.
        db_type = self.var_db_type.get()
        if db_type == "SQLite":
            if not self.var_sqlite_path.get().strip():
                messagebox.showerror("Fehler", "Bitte eine SQLite-Datenbankdatei auswählen.")
                return
        else:
            if not self.var_server.get().strip() or not self.var_database.get().strip():
                messagebox.showerror("Fehler", "Bitte Server und Datenbank angeben.")
                return

        # Aktuelle Eingaben in config.json speichern (außer Passwort – siehe oben).
        # So sind alle Felder beim nächsten Start bereits vorausgefüllt.
        self._config_aktualisieren(
            db_type=db_type,
            server=self.var_server.get().strip(),
            database=self.var_database.get().strip(),
            user=self.var_user.get().strip(),
            port=self.var_port.get().strip(),
            sqlite_path=self.var_sqlite_path.get().strip(),
            trusted=self.var_trusted.get(),
            input_file=input_file,
            output_dir=output_dir
            # Passwort wird bewusst NICHT in config.json gespeichert
        )

        # Button deaktivieren, damit der Benutzer nicht doppelt klickt.
        # NEU: Bei ctk.CTkButton lautet der Zustand "disabled" (wie bei tkinter).
        self.btn_run.configure(state="disabled")

        # Fortschrittsbalken starten (animiert, solange die Verarbeitung läuft)
        self.progress.start(12)

        # Verarbeitung in einem Hintergrund-Thread starten.
        # daemon=True → der Thread wird automatisch beendet, wenn das Hauptfenster geschlossen wird.
        threading.Thread(target=self._execute, args=(input_file, output_dir),
                         daemon=True).start()


    def _execute(self, input_file, output_dir):
        """
        Führt die eigentliche Arbeit aus – wird im Hintergrund-Thread aufgerufen.
        Ruft nacheinander alle Kernfunktionen auf und schreibt jeden Schritt in den Log.

        Wichtig: GUI-Updates (z.B. messagebox) dürfen im Thread nicht direkt aufgerufen werden.
        Stattdessen wird self.after(0, ...) verwendet, um sie sicher im Haupt-Thread auszuführen.
        """
        try:
            db_type = self.var_db_type.get()

            # ── Schritt 1: Datenbankverbindung herstellen ──
            # Je nach gewähltem Datenbanktyp wird die passende Verbindungsfunktion aufgerufen.
            # Alle vier Funktionen geben ein Connection-Objekt zurück, das pandas versteht.
            self.log_write(f"Verbinde mit {db_type} …", "info")

            if db_type == "Microsoft SQL Server":
                conn = verbinde_mssql(
                    server=self.var_server.get().strip(),
                    database=self.var_database.get().strip(),
                    user=self.var_user.get().strip(),
                    password=self.var_password.get(),
                    trusted=self.var_trusted.get()   # Windows-Auth oder SQL-Login
                )

            elif db_type == "SQLite":
                # SQLite braucht nur den Dateipfad – kein Benutzer, kein Passwort
                conn = verbinde_sqlite(self.var_sqlite_path.get().strip())

            elif db_type == "MySQL":
                conn = verbinde_mysql(
                    server=self.var_server.get().strip(),
                    database=self.var_database.get().strip(),
                    user=self.var_user.get().strip(),
                    password=self.var_password.get(),
                    port=self.var_port.get().strip() or 3306  # Fallback auf Standardport
                )
            elif db_type == "PostgreSQL":
                conn = verbinde_postgresql(
                    server=self.var_server.get().strip(),
                    database=self.var_database.get().strip(),
                    user=self.var_user.get().strip(),
                    password=self.var_password.get(),
                    port=self.var_port.get().strip() or 5432  # Fallback auf Standardport
                )
            else:
                raise RuntimeError(f"Unbekannter Datenbanktyp: {db_type}")

            self.log_write(f"✅ Verbindung zu {db_type} hergestellt.", "ok")

            # ── Schritt 2: SQL-Datei einlesen ──
            self.log_write(f"Lese SQL-Datei: {os.path.basename(input_file)}", "info")
            sql_query = lade_sql_datei(input_file)
            self.log_write("✅ SQL-Datei eingelesen.", "ok")

            # ── Schritt 3: SQL-Abfrage ausführen und Ergebnis als DataFrame speichern ──
            self.log_write("Führe SQL-Abfrage aus …", "info")
            df = sql_dataframe_erstellen(sql_query, conn)
            self.log_write(f"✅ Abfrage erfolgreich – {len(df):,} Zeilen, {len(df.columns)} Spalten.", "ok")

            # ── Schritt 4: Output-Dateinamen aus dem Input-Dateinamen ableiten ──
            # Beispiel: "Meine Abfrage.sql" → "Meine Abfrage_output.xlsx"
            base_name   = os.path.splitext(os.path.basename(input_file))[0]
            output_file = base_name + "_output.xlsx"

            # ── Schritt 5: DataFrame als Excel-Datei exportieren ──
            self.log_write(f"Exportiere nach: {output_file}", "info")
            full_path = export_to_excel(df, output_dir, output_file)
            self.log_write(f"✅ Export abgeschlossen: {full_path}", "ok")

            # Datenbankverbindung sauber schließen – guter Stil und verhindert
            # das Blockieren der Datenbank durch offene Verbindungen.
            conn.close()

            # Erfolgsmeldung im Haupt-Thread anzeigen (nicht direkt aus dem Thread!)
            self.after(0, lambda: messagebox.showinfo(
                "Fertig", f"Export erfolgreich abgeschlossen!\n\n{full_path}"))

        except Exception as e:
            # Bei jedem unerwarteten Fehler: Fehlermeldung in den Log schreiben
            # und einen Fehlerdialog im Haupt-Thread anzeigen
            error_msg = str(e)
            self.log_write(f"❌ Fehler: {e}", "err")
            self.after(0, lambda: messagebox.showerror("Fehler", error_msg))

        finally:
            # Dieser Block wird immer ausgeführt – egal ob Erfolg oder Fehler.
            # Fortschrittsbalken stoppen und Button wieder aktivieren.
            self.after(0, self.progress.stop)
            self.after(0, lambda: self.btn_run.configure(state="normal"))


# ─────────────────────────────────────────────────────────────────────────────
# Programm starten
# ─────────────────────────────────────────────────────────────────────────────

# Dieser Block wird nur ausgeführt, wenn die Datei direkt gestartet wird
# (nicht wenn sie als Modul in ein anderes Skript importiert wird).
if __name__ == "__main__":
    app = App()       # Fenster erzeugen
    app.mainloop()    # CustomTkinter-Ereignisschleife starten – hält das Fenster offen
